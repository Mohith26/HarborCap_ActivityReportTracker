"""Parser for leasing report style Excel files (Format B: 190 Industrial)."""

import re
import openpyxl
from app.parsers.base_parser import BaseParser, ParseResult
from app.parsers.normalizer import parse_date_flexible, parse_sf_range, parse_broker_contact


class ExcelLeasingReportParser(BaseParser):
    """Parses Excel files with property-level metrics and sectioned deal tables."""

    def can_parse(self, file_path: str, file_type: str, sheet_names: list[str] | None = None) -> bool:
        if file_type not in ("xlsx", "xlsm"):
            return False
        # This format is detected by checking file content, not sheet names
        # We check for "Leasing Report" patterns in the data
        try:
            wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
            ws = wb[wb.sheetnames[0]]
            # Scan first 20 rows for leasing report markers
            for row in ws.iter_rows(max_row=20, values_only=True):
                text = " ".join(str(cell) for cell in row if cell)
                if any(marker in text.upper() for marker in ["PROPERTY SF", "VACANT SF", "% VACANT", "LEASING REPORT"]):
                    wb.close()
                    return True
            wb.close()
        except Exception:
            pass
        return False

    def parse(self, file_path: str) -> ParseResult:
        result = ParseResult()
        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        ws = wb[wb.sheetnames[0]]
        result.sheet_names = wb.sheetnames

        rows = list(ws.iter_rows(values_only=True))
        wb.close()

        # Extract property-level metrics from the top section
        metrics = self._extract_metrics(rows)
        if metrics:
            result.metrics = metrics

        # Find and parse deal sections
        sections = self._find_sections(rows)
        for section_name, section_rows, stage in sections:
            deals = self._parse_section_deals(section_rows, stage)
            result.deals.extend(deals)

        result.raw_data["metrics"] = metrics
        result.raw_data["sections"] = [s[0] for s in sections]
        result.raw_data["total_deals"] = len(result.deals)

        return result

    def _extract_metrics(self, rows: list[tuple]) -> dict:
        """Extract property-level metrics from the top section."""
        metrics = {}
        for row in rows[:30]:  # Metrics are typically in first 30 rows
            if not row:
                continue
            for i, cell in enumerate(row):
                if cell is None:
                    continue
                cell_str = str(cell).strip().upper()

                if "PROPERTY SF" in cell_str or "TOTAL PROPERTY" in cell_str:
                    val = self._get_adjacent_value(row, i)
                    if val:
                        metrics["total_property_sf"] = self._parse_int(val)

                elif cell_str in ("VACANT SF", "TOTAL VACANT SF"):
                    val = self._get_adjacent_value(row, i)
                    if val:
                        metrics["vacant_sf"] = self._parse_int(val)

                elif "% VACANT" in cell_str:
                    val = self._get_adjacent_value(row, i)
                    if val:
                        metrics["vacancy_pct"] = self._parse_pct(val)

                elif "% OCCUPIED" in cell_str or "OCCUPANCY" in cell_str:
                    val = self._get_adjacent_value(row, i)
                    if val:
                        metrics["occupancy_pct"] = self._parse_pct(val)

                elif "INQUIR" in cell_str and "SENT" in cell_str:
                    val = self._get_adjacent_value(row, i)
                    if val:
                        metrics["inquiries_sent"] = self._parse_int(val)

                elif "TOUR" in cell_str and any(w in cell_str for w in ["CONDUCTED", "TOTAL", "COUNT"]):
                    val = self._get_adjacent_value(row, i)
                    if val:
                        metrics["tours_conducted"] = self._parse_int(val)

                elif "PROPOSAL" in cell_str and "SENT" in cell_str:
                    val = self._get_adjacent_value(row, i)
                    if val:
                        metrics["proposals_sent"] = self._parse_int(val)

        return metrics

    def _find_sections(self, rows: list[tuple]) -> list[tuple[str, list[tuple], str]]:
        """Find deal sections (SECTION I, II, etc.) and their row ranges."""
        sections = []
        section_patterns = [
            (r"SECTION\s*I[:\s].*(?:ACTIVE|PROPOSAL)", "2-LOI"),
            (r"SECTION\s*II[:\s].*(?:TOUR)", "3-Touring"),
            (r"SECTION\s*III[:\s].*(?:PROSPECT|TRACKING)", "4-Inquiry"),
            (r"SECTION\s*IV[:\s].*(?:DEAD|COMPLETE)", "7-Dead"),
            (r"SECTION\s*V[:\s].*(?:DEAD|INQUIR)", "7-Dead"),
        ]

        section_starts = []
        for row_idx, row in enumerate(rows):
            if not row:
                continue
            row_text = " ".join(str(cell) for cell in row if cell).upper()
            for pattern, stage in section_patterns:
                if re.search(pattern, row_text, re.IGNORECASE):
                    section_starts.append((row_idx, row_text.strip(), stage))
                    break

        # Build section row ranges
        for i, (start_idx, name, stage) in enumerate(section_starts):
            end_idx = section_starts[i + 1][0] if i + 1 < len(section_starts) else len(rows)
            section_rows = rows[start_idx:end_idx]
            sections.append((name, section_rows, stage))

        return sections

    def _parse_section_deals(self, section_rows: list[tuple], default_stage: str) -> list[dict]:
        """Parse deals from a section of rows."""
        deals = []
        if len(section_rows) < 2:
            return deals

        # Find header row within the section
        header_idx = None
        for idx, row in enumerate(section_rows):
            if not row:
                continue
            row_text = " ".join(str(cell).upper() for cell in row if cell)
            if "TENANT" in row_text or "NAME" in row_text:
                header_idx = idx
                break

        if header_idx is None:
            return deals

        header = section_rows[header_idx]

        # Map columns by position (these reports have varying column structures)
        col_map = {}
        for i, cell in enumerate(header):
            if cell is None:
                continue
            cell_str = str(cell).upper().strip()
            if "TENANT" in cell_str or "NAME" in cell_str:
                col_map["tenant"] = i
            elif "DATE" in cell_str:
                col_map["date"] = i
            elif "SF" in cell_str or "SIZE" in cell_str:
                col_map["size"] = i
            elif "RATE" in cell_str:
                col_map["rate"] = i
            elif "BROKER" in cell_str or "CO" in cell_str:
                col_map["broker"] = i
            elif "COMMENT" in cell_str or "FEEDBACK" in cell_str:
                col_map["comments"] = i

        for row_idx in range(header_idx + 1, len(section_rows)):
            row = section_rows[row_idx]
            if not row or all(cell is None or str(cell).strip() == "" for cell in row):
                continue

            tenant_idx = col_map.get("tenant")
            if tenant_idx is None or tenant_idx >= len(row) or not row[tenant_idx]:
                continue

            tenant = str(row[tenant_idx]).strip()
            if not tenant or tenant.upper() in ("PROPOSED TENANT NAME:", "TENANT NAME:", ""):
                continue

            size_raw = str(row[col_map["size"]]).strip() if col_map.get("size") is not None and col_map["size"] < len(row) and row[col_map["size"]] else None
            size_min, size_max = parse_sf_range(size_raw)

            broker_raw = str(row[col_map["broker"]]).strip() if col_map.get("broker") is not None and col_map["broker"] < len(row) and row[col_map["broker"]] else None
            broker_info = parse_broker_contact(broker_raw)

            comments = str(row[col_map["comments"]]).strip() if col_map.get("comments") is not None and col_map["comments"] < len(row) and row[col_map["comments"]] else None

            date_raw = row[col_map["date"]] if col_map.get("date") is not None and col_map["date"] < len(row) else None

            deal = {
                "stage": default_stage,
                "stage_numeric": int(default_stage.split("-")[0]) if "-" in default_stage else None,
                "deal_type": "New Deal",
                "tenant_name": tenant,
                "tenant_industry": None,
                "is_undisclosed": False,
                "broker_name": broker_info["name"],
                "broker_firm": broker_info["firm"],
                "broker_phone": broker_info["phone"],
                "broker_email": broker_info["email"],
                "initial_inquiry": parse_date_flexible(date_raw),
                "size_min_sf": size_min,
                "size_max_sf": size_max,
                "size_raw": size_raw,
                "commencement": None,
                "transaction_type": "Lease",
                "comments": comments,
                "last_updated": None,
                "last_updated_by": None,
                "lead_contact": None,
                "building_id": None,
            }
            deals.append(deal)

        return deals

    def _get_adjacent_value(self, row: tuple, idx: int):
        """Get the next non-None value in the row after idx."""
        for i in range(idx + 1, len(row)):
            if row[i] is not None:
                return row[i]
        return None

    def _parse_int(self, val) -> int | None:
        try:
            return int(float(str(val).replace(",", "").replace("%", "").strip()))
        except (ValueError, TypeError):
            return None

    def _parse_pct(self, val) -> float | None:
        try:
            v = float(str(val).replace("%", "").replace(",", "").strip())
            # If value is between 0 and 1, convert to percentage
            if 0 < v < 1:
                v = v * 100
            return round(v, 2)
        except (ValueError, TypeError):
            return None
