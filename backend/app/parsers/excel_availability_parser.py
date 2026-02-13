"""Parser for Excel 'Availabilities' sheets (Humble, Meadowedge)."""

import openpyxl
from app.parsers.base_parser import BaseParser, ParseResult
from app.parsers.column_mapper import map_columns, AVAILABILITY_ALIASES


class ExcelAvailabilityParser(BaseParser):
    """Parses 'Availabilities' sheets from Excel files."""

    def can_parse(self, file_path: str, file_type: str, sheet_names: list[str] | None = None) -> bool:
        if file_type not in ("xlsx", "xlsm"):
            return False
        if sheet_names:
            return any("availabilit" in s.lower() for s in sheet_names)
        return False

    def parse(self, file_path: str) -> ParseResult:
        result = ParseResult()
        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)

        for sheet_name in wb.sheetnames:
            if "availabilit" not in sheet_name.lower():
                continue

            ws = wb[sheet_name]
            rows = list(ws.iter_rows(values_only=True))
            if not rows:
                continue

            header_idx = self._find_header_row(rows)
            if header_idx is None:
                continue

            header = [str(cell) if cell else "" for cell in rows[header_idx]]
            mapping = map_columns(header, AVAILABILITY_ALIASES)

            if "building" not in mapping and "total_size_sf" not in mapping:
                continue

            for row_idx in range(header_idx + 1, len(rows)):
                row = rows[row_idx]
                if not row or all(cell is None or str(cell).strip() == "" for cell in row):
                    continue

                avail = self._parse_availability_row(row, mapping)
                if avail and (avail.get("building") or avail.get("total_size_sf")):
                    result.availabilities.append(avail)

            result.raw_data[f"availabilities_{sheet_name}"] = {
                "sheet": sheet_name,
                "header": header,
                "row_count": len(result.availabilities),
            }

        wb.close()
        return result

    def _find_header_row(self, rows: list[tuple], max_scan: int = 10) -> int | None:
        for idx in range(min(max_scan, len(rows))):
            row = rows[idx]
            if not row:
                continue
            row_text = " ".join(str(cell).upper() for cell in row if cell)
            if "BUILDING" in row_text or "TOTAL SIZE" in row_text or "STATUS" in row_text:
                return idx
        return None

    def _parse_availability_row(self, row: tuple, mapping: dict[str, int]) -> dict | None:

        def get(field: str):
            idx = mapping.get(field)
            if idx is None or idx >= len(row):
                return None
            return row[idx]

        def get_str(field: str) -> str | None:
            val = get(field)
            return str(val).strip() if val is not None else None

        def get_int(field: str) -> int | None:
            val = get(field)
            if val is None:
                return None
            try:
                return int(float(str(val).replace(",", "")))
            except (ValueError, TypeError):
                return None

        def get_decimal(field: str):
            val = get(field)
            if val is None:
                return None
            try:
                cleaned = str(val).replace("$", "").replace(",", "").strip()
                return float(cleaned) if cleaned else None
            except (ValueError, TypeError):
                return None

        building = get_str("building")
        if building and building.upper() in ("BUILDING", "BLDG"):
            return None  # Skip header-like rows

        return {
            "building": building,
            "total_size_sf": get_int("total_size_sf"),
            "status": get_str("status"),
            "office_sf": get_int("office_sf"),
            "lease_rate_psf": get_decimal("lease_rate_psf"),
            "opex": get_decimal("opex"),
            "sale_price_psf": get_decimal("sale_price_psf"),
            "power_amps": get_int("power_amps"),
            "loading_doors": get_str("loading_doors"),
            "eave_height": get_str("eave_height"),
            "crane_ready": get_str("crane_ready"),
            "land_acreage": get_decimal("land_acreage"),
            "additional_info": get_str("additional_info"),
        }
