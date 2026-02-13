"""Parser for standard Excel deal sheets (Format A: Humble, Meadowedge, Northport)."""

import openpyxl
from app.parsers.base_parser import BaseParser, ParseResult
from app.parsers.column_mapper import map_columns, COLUMN_ALIASES
from app.parsers.normalizer import (
    parse_sf_range, parse_stage, parse_broker_contact,
    parse_date_flexible, parse_transaction_type, is_undisclosed_tenant,
)


class ExcelDealParser(BaseParser):
    """Parses Excel files with a 'Deal Name' sheet or similar deal tracking structure."""

    # Sheet names that indicate deal tracking data
    DEAL_SHEET_NAMES = {"Deal Name", "Deal Sheet", "Deals", "Sheet1"}

    def can_parse(self, file_path: str, file_type: str, sheet_names: list[str] | None = None) -> bool:
        if file_type not in ("xlsx", "xlsm"):
            return False
        if sheet_names:
            return any(s in self.DEAL_SHEET_NAMES for s in sheet_names)
        # If no sheet names provided, try to open and check
        return True  # We'll handle detection in parse()

    def parse(self, file_path: str) -> ParseResult:
        result = ParseResult()
        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        result.sheet_names = wb.sheetnames

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows = list(ws.iter_rows(values_only=True))
            if not rows:
                continue

            # Find the header row (contains "STAGE" or "DEAL TYPE")
            header_idx = self._find_header_row(rows)
            if header_idx is None:
                continue

            header = [str(cell) if cell else "" for cell in rows[header_idx]]
            mapping = map_columns(header, COLUMN_ALIASES)

            if "stage" not in mapping and "tenant_name" not in mapping:
                continue

            # Parse data rows
            for row_idx in range(header_idx + 1, len(rows)):
                row = rows[row_idx]
                if not row or all(cell is None or str(cell).strip() == "" for cell in row):
                    continue

                deal = self._parse_deal_row(row, mapping)
                if deal and deal.get("tenant_name"):
                    result.deals.append(deal)

            raw_key = f"deals_{sheet_name}"
            result.raw_data[raw_key] = {
                "sheet": sheet_name,
                "header": header,
                "row_count": len(rows) - header_idx - 1,
            }

        wb.close()
        return result

    def _find_header_row(self, rows: list[tuple], max_scan: int = 10) -> int | None:
        """Find the row containing column headers."""
        for idx in range(min(max_scan, len(rows))):
            row = rows[idx]
            if not row:
                continue
            row_text = " ".join(str(cell).upper() for cell in row if cell)
            if "STAGE" in row_text or "DEAL TYPE" in row_text or "TENANT" in row_text:
                return idx
        return None

    def _parse_deal_row(self, row: tuple, mapping: dict[str, int]) -> dict | None:
        """Parse a single data row into a normalized deal dict."""

        def get(field: str) -> str | None:
            idx = mapping.get(field)
            if idx is None or idx >= len(row):
                return None
            val = row[idx]
            return str(val).strip() if val is not None else None

        stage_raw = get("stage")
        tenant_raw = get("tenant_name")

        # Skip rows that are section headers or empty
        if not tenant_raw and not stage_raw:
            return None
        if tenant_raw and tenant_raw.upper() in ("TENANT / INDUSTRY", "TENANT / INDUSTRY / BUYER"):
            return None

        stage, stage_numeric = parse_stage(stage_raw)
        size_min, size_max = parse_sf_range(get("size"))
        broker_info = parse_broker_contact(get("broker_contact"))
        initial_inquiry = parse_date_flexible(row[mapping["initial_inquiry"]] if "initial_inquiry" in mapping and mapping["initial_inquiry"] < len(row) else None)
        last_updated = parse_date_flexible(row[mapping["last_updated"]] if "last_updated" in mapping and mapping["last_updated"] < len(row) else None)

        return {
            "stage": stage,
            "stage_numeric": stage_numeric,
            "deal_type": get("deal_type"),
            "tenant_name": tenant_raw,
            "tenant_industry": None,  # Often same column as tenant_name
            "is_undisclosed": is_undisclosed_tenant(tenant_raw),
            "broker_name": broker_info["name"],
            "broker_firm": broker_info["firm"],
            "broker_phone": broker_info["phone"],
            "broker_email": broker_info["email"],
            "initial_inquiry": initial_inquiry,
            "size_min_sf": size_min,
            "size_max_sf": size_max,
            "size_raw": get("size"),
            "commencement": get("commencement"),
            "transaction_type": parse_transaction_type(get("transaction_type")),
            "comments": get("comments"),
            "last_updated": last_updated,
            "last_updated_by": get("last_updated_by"),
            "lead_contact": get("lead_contact"),
            "building_id": None,
        }
