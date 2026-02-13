"""Parser for per-building Excel deal sheets (Format C: Corbin Road Business Park)."""

import openpyxl
from app.parsers.base_parser import BaseParser, ParseResult
from app.parsers.column_mapper import map_columns, COLUMN_ALIASES
from app.parsers.normalizer import (
    parse_sf_range, parse_stage, parse_broker_contact,
    parse_date_flexible, parse_transaction_type, is_undisclosed_tenant,
)


class ExcelBuildingSheetParser(BaseParser):
    """Parses Excel files with multiple sheets, one per building (e.g., Corbin Road)."""

    # Sheets to skip
    SKIP_SHEETS = {"Investor List", "Summary", "Availabilities", "Deal Name", "Sheet1"}

    def can_parse(self, file_path: str, file_type: str, sheet_names: list[str] | None = None) -> bool:
        if file_type not in ("xlsx", "xlsm"):
            return False
        if not sheet_names:
            return False
        # If there are many sheets (>3) and they don't match standard patterns, likely per-building
        non_standard = [s for s in sheet_names if s not in self.SKIP_SHEETS]
        return len(non_standard) >= 3

    def parse(self, file_path: str) -> ParseResult:
        result = ParseResult()
        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        result.sheet_names = wb.sheetnames

        for sheet_name in wb.sheetnames:
            if sheet_name in self.SKIP_SHEETS:
                continue

            ws = wb[sheet_name]
            rows = list(ws.iter_rows(values_only=True))
            if not rows:
                continue

            header_idx = self._find_header_row(rows)
            if header_idx is None:
                continue

            header = [str(cell) if cell else "" for cell in rows[header_idx]]
            mapping = map_columns(header, COLUMN_ALIASES)

            if "stage" not in mapping and "tenant_name" not in mapping:
                continue

            building_id = sheet_name

            for row_idx in range(header_idx + 1, len(rows)):
                row = rows[row_idx]
                if not row or all(cell is None or str(cell).strip() == "" for cell in row):
                    continue

                deal = self._parse_deal_row(row, mapping, building_id)
                if deal and deal.get("tenant_name"):
                    result.deals.append(deal)

            result.raw_data[f"building_{sheet_name}"] = {
                "sheet": sheet_name,
                "building_id": building_id,
                "header": header,
                "deal_count": sum(1 for d in result.deals if d.get("building_id") == building_id),
            }

        wb.close()
        return result

    def _find_header_row(self, rows: list[tuple], max_scan: int = 10) -> int | None:
        for idx in range(min(max_scan, len(rows))):
            row = rows[idx]
            if not row:
                continue
            row_text = " ".join(str(cell).upper() for cell in row if cell)
            if "STAGE" in row_text or "DEAL TYPE" in row_text or "TENANT" in row_text:
                return idx
        return None

    def _parse_deal_row(self, row: tuple, mapping: dict[str, int], building_id: str) -> dict | None:

        def get(field: str) -> str | None:
            idx = mapping.get(field)
            if idx is None or idx >= len(row):
                return None
            val = row[idx]
            return str(val).strip() if val is not None else None

        stage_raw = get("stage")
        tenant_raw = get("tenant_name")

        if not tenant_raw and not stage_raw:
            return None
        if tenant_raw and tenant_raw.upper() in ("TENANT / INDUSTRY", "TENANT / INDUSTRY / BUYER"):
            return None

        stage, stage_numeric = parse_stage(stage_raw)
        size_min, size_max = parse_sf_range(get("size"))
        broker_info = parse_broker_contact(get("broker_contact"))
        initial_inquiry = parse_date_flexible(
            row[mapping["initial_inquiry"]] if "initial_inquiry" in mapping and mapping["initial_inquiry"] < len(row) else None
        )

        return {
            "stage": stage,
            "stage_numeric": stage_numeric,
            "deal_type": get("deal_type"),
            "tenant_name": tenant_raw,
            "tenant_industry": None,
            "is_undisclosed": is_undisclosed_tenant(tenant_raw),
            "broker_name": broker_info["name"],
            "broker_firm": broker_info["firm"],
            "broker_phone": broker_info["phone"],
            "broker_email": broker_info["email"],
            "initial_inquiry": initial_inquiry,
            "size_min_sf": size_min,
            "size_max_sf": size_max,
            "size_raw": get("size"),
            "commencement": get("commencement"),
            "transaction_type": parse_transaction_type(get("transaction_type")),
            "comments": get("comments"),
            "last_updated": None,
            "last_updated_by": None,
            "lead_contact": None,
            "building_id": building_id,
        }
